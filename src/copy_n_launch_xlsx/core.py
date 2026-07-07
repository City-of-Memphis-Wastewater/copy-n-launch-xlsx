# src/copy_n_launch_xlsx/core.py

from __future__ import annotations
from pathlib import Path
import shutil
import openpyxl
import logging
from typing import Optional
from dataclasses import dataclass
import pyhabitat
from datetime import date, timedelta
from openpyxl.workbook.defined_name import DefinedName
import sys

from .paths import BLANK_DAILY_XLSX, get_target_copy_dir

FILENAME_FORMAT = "daily_%Y-%m-%d.xlsx"
logger = logging.getLogger(__name__)

@dataclass
class CopyResult:
    destination: Optional[str] = None
    is_new: Optional[bool] = False

    @property
    def status_message(self) -> str:
        """Statuses."""
        return {
            True: "File copied.",
            False: "File exists.",
            None: "Exited."
        }.get(self.is_new, "Error.")

    def __bool__(self):
        return self.destination is not None
    

def build_filename(day: date | None = None) -> str:
    if day is None:
        day = date.today()
    filename = day.strftime(FILENAME_FORMAT)
    return filename


def copy_then_launch(day: date | None = None) -> CopyResult:
    """Copies the blank template and launches it for a specific day (defaults to today)."""
    if day is None:
        day = date.today()

    target_dir = get_target_copy_dir()
    target_dir.mkdir(parents=True, exist_ok=True)

    destination = target_dir / build_filename(day=day)
    logger.debug(f"{destination=}")
    # Check if the file is already there
    if destination.exists():
        logger.info(f"Daily file already exists at {destination}. Skipping copy. Launching existing file.")
        pyhabitat.launch_file(destination)
        #return destination
        return CopyResult(destination=destination,is_new=False)
    if BLANK_DAILY_XLSX.exists():
        shutil.copy2(BLANK_DAILY_XLSX, destination)
    else:
        raise FileNotFoundError(
            f"Not found: Expected blank spreadsheet at {BLANK_DAILY_XLSX}"
        )
        #print(f"Please put daily_blank.xlsx in the expected place: {BLANK_DAILY_XLSX}")
        #sys.exit(0)
    # Open/save if you later want to update named ranges,
    # dates, workbook properties, etc.
    wb = openpyxl.load_workbook(destination)

    # future edits go here
    set_date_in_spreadsheet(wb, day)
    wb.save(destination)

    pyhabitat.launch_file(destination)

    return CopyResult(destination=destination,is_new=True)


def set_date_in_spreadsheet(wb, day: date | None = None):
    if day is None:
        day = date.today()
    day_str = day.strftime("%m/%d/%Y")
    
    if "date" in wb.defined_names:
        defn = wb.defined_names["date"]
        
        # Use whichever field contains the reference string
        raw_value = defn.value if defn.value else defn.attr_text
        
        if raw_value and "!" in raw_value:
            sheet_name, cell_coord = raw_value.split("!")
            sheet_name = sheet_name.strip("'")      # Strip potential quote wrapping
            cell_coord = cell_coord.replace("$", "") # Strip absolute reference anchoring
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws[cell_coord] = day_str
                logger.debug(f"Successfully updated cell {cell_coord} on sheet '{sheet_name}' to {day_str}.")
        else:
            defn.value = f'="{day_str}"'
            logger.debug(f"Updated global value expression for named range 'date' to {day_str}.")
    else:
        new_name = DefinedName("date", attr_text=f'="{day_str}"')
        wb.defined_names.add(new_name)
        logger.debug(f"Named range 'date' not found. Created global expression variable set to {day_str}.")

def launch_tomorrow() -> CopyResult:
    """Dedicated function to build/launch the file for the day after today."""
    tomorrow = date.today() + timedelta(days=1)
    logger.info(f"Targeting tomorrow's file: {tomorrow}")
    return copy_then_launch(day=tomorrow)


def launch_yesterday_if_exists() -> CopyResult | None:
    """Launches yesterday's file ONLY if it already exists. Does not create it."""
    yesterday = date.today() - timedelta(days=1)
    target_dir = get_target_copy_dir()
    destination = target_dir / build_filename(yesterday)
    
    if destination.exists():
        logger.info(f"Yesterday's file found at {destination}. Launching.")
        pyhabitat.launch_file(destination)
        return CopyResult(destination=destination, is_new=False)
    
    logger.warning(f"Yesterday's file ({destination}) does not exist. Skipping launch.")
    return None